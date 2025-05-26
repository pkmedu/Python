# -*- coding: utf-8 -*-
"""
Created on Mon May  5 12:32:20 2025

@author: muhuri
"""

# -*- coding: utf-8 -*-
"""
Created on Mon May  5 10:46:04 2025
@author: muhuri
"""

# %pip install openpyxl
import pandas as pd
import glob
import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(message)s')

# Set paths
folder_path = 'C:/Links'
combined_csv_path = 'C:/Data/combined_all.csv'
final_excel_path = 'C:/Data/Combined_all.xlsx'

# Find all CSV files in the folder
csv_files = glob.glob(os.path.join(folder_path, '*.csv'))

if not csv_files:
    raise FileNotFoundError(f"No CSV files found in {folder_path}")

# Read and combine CSVs
df_list = [pd.read_csv(file) for file in csv_files]
combined_df = pd.concat(df_list, ignore_index=True)

# Save combined data as CSV (optional)
combined_df.to_csv(combined_csv_path, index=False)

# Reload to ensure consistency
df = pd.read_csv(combined_csv_path)

# Count rows before and after deduplication
initial_count = len(df)

# Ensure 'Link' column exists
if 'Link' not in df.columns:
    raise ValueError("Column 'Link' not found in the CSV files.")

# Drop duplicates based on 'Link'
df_unique = df.drop_duplicates(subset='Link', keep='first')
final_count = len(df_unique)
duplicates_removed = initial_count - final_count

# Sort by 'Link'
df_unique_sorted = df_unique.sort_values(by='Link')

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = 'Links'

# Write header row
ws.append(list(df_unique_sorted.columns))

# Write data rows with hyperlinks directly embedded
for row in df_unique_sorted.itertuples(index=False):
    ws_row = ws.max_row + 1
    for col_idx, value in enumerate(row):
        cell = ws.cell(row=ws_row, column=col_idx + 1)
        col_name = df_unique_sorted.columns[col_idx]

        if col_name == 'Link' and pd.notnull(value):
            cell.value = value
            cell.hyperlink = value
            cell.font = Font(color='0000FF', underline='single')  # Blue, underlined font
        else:
            cell.value = value

# Adjust column widths
for col_idx, column in enumerate(df_unique_sorted.columns, 1):
    max_length = max((len(str(cell)) for cell in df_unique_sorted[column]), default=0)
    adjusted_width = min(max_length + 2, 100)  # Cap width at 100 to prevent extremes
    ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

# Save Excel file
wb.save(final_excel_path)

# Log summary
logging.info(f"Total rows before deduplication: {initial_count}")
logging.info(f"Total rows after deduplication: {final_count}")
logging.info(f"Duplicates removed: {duplicates_removed}")
logging.info(f"Deduplicated and sorted Excel file with clickable hyperlinks saved to: {final_excel_path}")
