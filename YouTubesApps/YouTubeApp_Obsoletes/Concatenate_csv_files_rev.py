# -*- coding: utf-8 -*-
"""
Created on Tue Apr 29 09:43:45 2025

@author: muhuri
"""
# %pip install openpyxl
import pandas as pd
import glob
import os
from openpyxl import Workbook
from openpyxl.styles import Font

# Set paths
folder_path = 'C:/Links'
combined_csv_path = 'C:/Data/combined_all.csv'
final_excel_path = 'C:/Data/Combined_all.xlsx'

# Find all CSV files in the folder
csv_files = glob.glob(os.path.join(folder_path, '*.csv'))

# Read and combine CSVs
df_list = [pd.read_csv(file) for file in csv_files]
combined_df = pd.concat(df_list, ignore_index=True)

# Save combined data as CSV (optional step)
combined_df.to_csv(combined_csv_path, index=False)

# Reload to ensure consistency
df = pd.read_csv(combined_csv_path)

# Count rows before and after deduplication
initial_count = len(df)

# Ensure 'Link' column exists
if 'Link' not in df.columns:
    raise ValueError("Column 'Link' not found in the CSV files.")

# Drop duplicates based on 'Link'
df_unique = df.drop_duplicates(subset='Link', keep='first')
final_count = len(df_unique)
duplicates_removed = initial_count - final_count

# Sort by 'Link'
df_unique_sorted = df_unique.sort_values(by='Link')

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = 'Links'

# Write header row
ws.append(list(df_unique_sorted.columns))

# Write data rows with embedded hyperlinks
for row in df_unique_sorted.itertuples(index=False):
    ws_row = ws.max_row + 1
    for col_idx, value in enumerate(row):
        cell = ws.cell(row=ws_row, column=col_idx + 1)
        col_name = df_unique_sorted.columns[col_idx]

        if col_name == 'Link' and pd.notnull(value):
            cell.value = value
            cell.hyperlink = value  # Embed hyperlink
            cell.font = Font(color='0000FF', underline='single')  # Blue, underlined
        else:
            cell.value = value

# Save Excel file
wb.save(final_excel_path)

# Print summary
print(f"Total rows before deduplication: {initial_count}")
print(f"Total rows after deduplication: {final_count}")
print(f"Duplicates removed: {duplicates_removed}")
print(f"Deduplicated and sorted Excel file with embedded hyperlinks saved to: {final_excel_path}")
